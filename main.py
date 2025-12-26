from fastapi import FastAPI,Request, WebSocket, WebSocketDisconnect, UploadFile, File
from fastapi.responses import PlainTextResponse, JSONResponse
from openai import OpenAI
from dotenv import load_dotenv
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import docx  # python-docx for Word files
import openpyxl
from openpyxl.utils import get_column_letter
import os
import io
from json_repair import repair_json
import re
# import whisper
import tempfile
import shutil
import ast
import uuid
from typing import List

from pydantic import BaseModel
import json 
import fitz  # PyMuPDF
from models import check_connection, push_resume_report,get_report, push_questions, get_questions, get_uuid

load_dotenv() 
# Load Whisper model
# model = whisper.load_model("base")  # or 'tiny', 'small', etc.

# OpenAI API Key (Set this in environment variables)
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

app = FastAPI()

# Enable CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://snaphr.vercel.app",],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)

dummy_excel = []
ScoreData = {}
count_followup = 0

class QuestionIDRequest(BaseModel):
    questions_id: str

class BulkResumeRequest(BaseModel):
    resumes: List[str]
    jd: str

class ResumeRequest(BaseModel):
    text: str
    jd: str

class Question(BaseModel):
    text: str

class QuestionAnswer(BaseModel):
    question: str
    answer: str

class CoverLetter(BaseModel):
    resume: str
    cover_letter: str

class ResumeReport(BaseModel):
    report: list
    repo_name: str

class ResumeReportName(BaseModel):
    repo_name: str
    
questions_list = []

# @app.websocket("/ws/audio")
# async def websocket_endpoint(websocket: WebSocket):
#     await websocket.accept()
#     try:
#         while True:
#             audio_chunk = await websocket.receive_bytes()

#             # Save to temp file
#             with tempfile.NamedTemporaryFile(delete=False, suffix=".wav") as tmp:
#                 tmp.write(audio_chunk)
#                 tmp_path = tmp.name

#             # Transcribe
#             result = model.transcribe(tmp_path)
#             await websocket.send_text(result["text"])
#     except WebSocketDisconnect:
#         print("Client disconnected")


def extract_text_from_pdf(file_bytes):
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    return "\n".join([page.get_text() for page in doc])

def extract_text_from_docx(file_bytes):
    doc = docx.Document(io.BytesIO(file_bytes))
    return "\n".join([para.text for para in doc.paragraphs])

@app.get("/download-excel/")
def download_excel():
    filename = "resumes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = [
        "name", "score", "phone", "email", "LinkedIn", "location", "total_experience",
        "skills", "summary"
    ]
    ws.append(headers)

    for item in dummy_excel:
        row = []
        for key in headers:
            val = item.get(key)
            if key=="score" and isinstance(val, str):
                print("type val")
                val = int(val[0])
            if isinstance(val, list):
                val = ", ".join(val)
            row.append(val)
        ws.append(row)

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)
    dummy_excel.clear()
    return FileResponse(path=filename, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get("/check-conection")
async def check_conn():
    res = await check_connection()
    return res

@app.post("/push-report")
async def push_repo(data:ResumeReport):
    collection_data = {}
    collection_data["report_name"] = data.repo_name
    collection_data["report_data"] = data.report
    res = await push_resume_report(collection_data)
    return res

@app.post("/push-questions")
async def push_quest(questions=questions_list):
    question_data = {}
    unique_id = str(uuid.uuid4())
    question_data["ID"] = unique_id
    question_data["questions"] = questions
    res = await push_questions(question_data)
    return unique_id

@app.post("/get-questions")
async def get_quest(res:QuestionIDRequest):
    result = await get_questions(res.questions_id)
    return result

@app.get("/get-uuid")
async def get_uuids():
    result = await get_uuid()
    return result

@app.post("/get-report")
async def get_repo(data:ResumeReportName):
    res = await get_report(data.repo_name)
    return res

@app.post("/upload-bulk-resume/")
async def upload_bulk_resume(files: list[UploadFile] = File(...)):
    extracted_texts = []
    dummy_excel = []

    for file in files:
        try:
            contents = await file.read()
            filename = file.filename.lower()

            if filename.endswith(".pdf"):
                text = extract_text_from_pdf(contents)
            elif filename.endswith(".docx"):
                text = extract_text_from_docx(contents)
            else:
                text = f"Unsupported file type: {filename}"

            extracted_texts.append({
                "filename": filename,
                "text": text
            })
        except:
            raise Exception("An error occured parsing the resume")

    return {"results": extracted_texts}

@app.post("/follow-up/", response_class=PlainTextResponse)
async def followUpQuestion(data: Question):
    prompt = f'''You are a interviewer for a technical interview round, 
                ask a followup question for this question: {data.text}. 
                Donot ask about where have you implemented it before, just ask
                a harder purely technical followup question.
                '''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    followup_question = completion.choices[0].message.content
    print("Follow Up,", followup_question)
    return followup_question

@app.post("/evaluate/", response_class=PlainTextResponse)
async def evaluateAnswer(data: QuestionAnswer):
    prompt = f'''You are a interviewer for a technical interview round, 
                evaluate the answer: {data.answer}, for the question: {data.question}.
                And give a score out of 10.
                
                Your Response should only STRICTLY be an Integer and nothing else.
                '''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    score = completion.choices[0].message.content
    
    
    ScoreData[data.question] = int(score)
    print("Score", ScoreData)
    return str(score)

@app.post("/answer/", response_class=JSONResponse)
async def exrtact_info(data: Question):
    prompt = f'''
                You are an assistant that always responds in JSON array format.

                Rules:
                1. If the question asks for code, a script, or implementation (e.g., "Write a Python function", "Create a script", "Implement..."), respond with the **full code** as a single string in a JSON array like: ["<entire code here>"] With a prefix "CODE".
                2. For all other types of questions (theory, conceptual, explanation), respond in **very brief**, **pointwise** manner in a JSON array of short strings like: ["Point 1", "Point 2", "Point 3"].
                3. Do not include any explanation or markdown formatting.

                Question: "{data.text}"

                Examples:
                Q: "Explain what is an API"
                A: ["APIs allow communication between software systems", "They define endpoints and data exchange formats"]

                Q: "Write a Python function to calculate factorial"
                A: ["CODEdef factorial(n):\\n    return 1 if n == 0 else n * factorial(n - 1)"]
                '''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    answer = completion.choices[0].message.content
    parsed_answer = json.loads(answer)
    print("Answer,", parsed_answer)
    return parsed_answer

@app.post("/extract-info-bulk/")
async def extract_info_bulk(data: BulkResumeRequest):
    prompt = f'''
    You are an AI that analyzes resumes against a job description.

    Job Description:
    "{data.jd}"

    For each resume provided, return a JSON array where each element corresponds to one resume with keys:
    [name, score, phone, email, LinkedIn, location, total_experience, skills, summary].

    Rules:
    - "score" is a rating out of 10 based on how well the resume matches the job description.
    - "summary" is a brief analysis.
    - for "skills" extract maximum of top 8 skills, not more than that.
    - Return ONLY valid JSON, no extra text.
    - Keep results in the same order as the resumes are given.
    - donot use any new line quotes like "/n" with key values

    Return STRICTLY STRICTLY ONLY a VALID JSON, with NO extra text, no markdown, no ```json fences..
    Do not include markdown, code blocks, or explanatory text.
    Return a JSON array directly, nothing else.
    '''

    for idx, resume in enumerate(data.resumes, start=1):
        prompt += f"\nResume {idx}:\n{resume}\n"

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )

    raw_output = completion.choices[0].message.content.strip()

    # Step 1: Clean markdown fences
    cleaned = re.sub(r"^```(json)?|```$", "", raw_output.strip(), flags=re.MULTILINE).strip()

    # Step 2: Try parsing, fallback to repair
    try:
        parsed = json.loads(cleaned)
    except json.JSONDecodeError:
        repaired = repair_json(cleaned)
        parsed = json.loads(repaired)

    dummy_excel.extend(parsed)
    return parsed


@app.post("/extract-info/")
async def exrtact_info(data: ResumeRequest):
    prompt = f'''Exrtact information from this resume and generate a valid json response, donot use any new line quotes like "/n" with
        key values: [name, score, phone, email, LinkedIn, location, total_experience, skills, summary].
        Also based on the job description: "{data.jd}" compare and analyze it with the resume of the candidate and give a rating out of 10 and put it in the "score" key.
        And give a brief summary about your analysis on the candidate and put it in "summary" key.
        This is the resume: {data.text}
    
        Return STRICTLY STRICTLY ONLY a VALID JSON, with NO extra text.'''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    improved_text = completion.choices[0].message.content
    dummy_excel.append(json.loads(improved_text))
    return json.loads(improved_text)


@app.post("/bot-questioner/")
async def bot_questioner(data: ResumeRequest):
    prompt = f'''Extract information from this resume and generate a valid JSON response without any new line quotes ("/n").
    The JSON structure should have the following keys: ["Introduction_question", "Skills_questions", "HLD_questions", "Experience_questions", "Coding_Question"].
    
    - The "Introduction_question" key should have only one brief intoduction of the candidate question like: Can you give a brief Introduction about your experience? .
    - The "Skills_questions" key should be a dictionary where each skill (e.g., "Python", "React", "Data Analysis") is a key and the value is a list of 3 purely technical questions related to that specific skill.
    - The "HLD_questions" key should contain 3 technical questions related to high-level design situations that a developer might encounter based on the candidate's resume.
    - The "Experience_questions" key should contain 3 questions based on the candidate's past projects and work experience.
    - The "Coding_Question" key should contain 3 python based coding questions and their with increasing level of difficulty.

    Analyze the resume content and the job description to formulate relevant questions, the "Skills_questions" should have all the skills mentioned in the job description.

    Job Description: "{data.jd}"
    Resume: {data.text}
    
    Return STRICTLY STRICTLY ONLY a VALID JSON, with NO extra text.'''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    questions = completion.choices[0].message.content
    print("Questions",json.loads(questions))
    return json.loads(questions)




@app.post("/ask-questions/")
async def exrtact_info(data: ResumeRequest):
    prompt = f'''Extract information from this resume and generate a valid JSON response without any new line quotes ("/n").
    The JSON structure should have the following keys: ["Skills_questions", "HLD_questions", "Experience_questions"].
    
    - The "Skills_questions" key should be a dictionary where each skill (e.g., "Python", "React", "Data Analysis") is a key and the value is a list of 3 purely technical questions related to that specific skill.
    - The "HLD_questions" key should contain 3 technical questions related to high-level design situations that a developer might encounter based on the candidate's resume.
    - The "Experience_questions" key should contain 3 questions based on the candidate's past projects and work experience.
    - The "Coding_Question" key should contain 5 code writing questions with mediium to hard difficulty level.

    Analyze the resume content and the job description to formulate relevant questions, the "Skills_questions" should have all the skills mentioned in the job description.

    Job Description: "{data.jd}"
    Resume: {data.text}
    
    Return STRICTLY STRICTLY ONLY a VALID JSON, with NO extra text.'''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    questions = completion.choices[0].message.content
    
    # Convert to dictionary
    data_dict = ast.literal_eval(questions)

    # Now you can iterate and append to a list


    for category, q in data_dict.items():
        if isinstance(q, dict):  # Skills_questions
            for subcat, qs in q.items():
                questions_list.extend(qs)
        else:  # HLD_questions, Experience_questions
            questions_list.extend(q)
    
    print("questions list: ", questions_list)

    return json.loads(questions)

@app.post("/upload-resume/")
async def upload_resume(file: UploadFile = File(...)):
    file_extension = file.filename.split(".")[-1].lower()
    contents = await file.read()

    if file_extension == "pdf":
        extracted_text = extract_text_from_pdf(contents)
    elif file_extension == "docx":
        extracted_text = extract_text_from_docx(contents)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Only PDF and DOCX are allowed.")

    return {"extracted_text": extracted_text}


@app.post("/improve-resume/")
async def improve_resume(data: ResumeRequest):
    prompt = f'''Improve this resume to be ats friendly and generate a valid json response, donot use any new line quotes like "/n") with
    key values: [name, summary, experience[jobTitle, company, location, duration, responsibilities], skills, education, certification, achivements, email, phone, address ]. keep these points in mind(
    Header: Your name, phone, email, LinkedIn, and location (City, State).
    Professional Summary: A 2-3 sentence summary highlighting your experience, skills, and achievements.
    Work Experience:
        Use bullet points to describe your impact in each role.
        Start each bullet with action verbs (e.g., "Developed," "Optimized," "Led").
        Do not repeat action words, insted use synonyms.
        Do not make grammer mistakes.
        Include metrics where possible (e.g., "Increased sales by 30% in 6 months").
        Note: If metrics is not mentioned in the resume then create your own metrics.
        Skills Section: List technical and job-relevant skills in a separate section.
        Education & Certifications: Mention degree, university, and year.),
        Make sure the resume is not too short, if it is then increase the lenth of resume.Specially increase the bullet points for each work experience, include day to day role and more.
        And Finally format the resume beautifully.
        This is the resume: {data.text}
        
        Return STRICTLY STRICTLY ONLY a VALID JSON, with NO extra text.'''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    

    improved_text = completion.choices[0].message.content
    # json_response = json.loads(improved_text)
    # print(json_response)
    return json.loads(improved_text)


@app.post("/cover-letter/")
async def coverletter(data: CoverLetter):
    prompt = f'''Tailor this resume according to the cover letter given and generate a json format response with
    key values: [name, summary, experience, skills, education, certification, achivements, email, phone, address ]. It should be ats friendly and keep these points in mind(
    Header: Your name, phone, email, LinkedIn, and location (City, State).
    Professional Summary: A 2-3 sentence summary highlighting your experience, skills, and achievements.
    Work Experience:
        Use bullet points to describe your impact in each role.
        Start each bullet with action verbs (e.g., "Developed," "Optimized," "Led").
        Do not repeat action words, insted use synonyms.
        Do not make grammer mistakes.
        Include metrics where possible (e.g., "Increased sales by 30% in 6 months").
        Note: If metrics is not mentioned in the resume then create your own metrics.
        Skills Section: List technical and job-relevant skills in a separate section.
        Education & Certifications: Mention degree, university, and year.),
        Make sure the resume is not too short, if it is then increase the lenth of resume.Specially increase the bullet points for each work experience, include day to day role and more.
        And Finally format the resume beautifully. 
        This is the resume: {data.resume}
        This is the cover letter: {data.cover_letter} 

        Return ONLY a VALID PYTHON DICTIONARY, with NO extra text. and START the string with triple quotes so that you don't have
        to use \" and please DONOT use \n anywhere'''

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": prompt}
        ]
    )
    
        # Parse the JSON response from OpenAI
    try:
        improved_text = json.loads(completion.choices[0].message.content)
    except json.JSONDecodeError:
        return {"error": "Invalid JSON format received from OpenAI."}

    return improved_text


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
